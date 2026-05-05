from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .models import ExtractionModels
from .text_utils import clean_text, format_number, normalize_for_match, numeric_or_none, parse_excel_date


HEADER_KEYWORDS = re.compile(
    r"\b(soll|ist|erfüllt|erfuellt|vertrag|endbericht|berichtigt|begründung|begruendung|anmerkung)\b",
    re.I,
)
REPORT_KEYWORDS = re.compile(
    r"(indikatorenbericht|endbericht indikatoren|zielindikatoren|soll|ist|erfüllt|erfuellt)",
    re.I,
)
INDICATOR_HINTS = re.compile(
    r"(anzahl|teilnehmer|teilnehm|projektteilnehm|veranstaltung|beratung|kurs|workshop|stunden|standorte)",
    re.I,
)


@dataclass
class Cell:
    row: int
    col: int
    value: object

    @property
    def text(self) -> str:
        return clean_text(self.value)

    @property
    def coordinate(self) -> str:
        return f"{get_column_letter(self.col)}{self.row}"


class SheetGrid:
    def __init__(self, title: str, rows: list[list[object]]):
        self.title = title
        self.rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(row) for row in rows), default=0)

    def value(self, row: int, col: int):
        if row < 1 or col < 1:
            return None
        if row > self.max_row:
            return None
        row_values = self.rows[row - 1]
        if col > len(row_values):
            return None
        return row_values[col - 1]

    def cell(self, row: int, col: int) -> Cell:
        return Cell(row=row, col=col, value=self.value(row, col))

    def row_values(self, row: int, max_col: int | None = None) -> list[str]:
        max_col = min(max_col or self.max_column, self.max_column)
        return [clean_text(self.value(row, col)) for col in range(1, max_col + 1)]


def load_workbook_grids(path: Path) -> list[SheetGrid]:
    path = Path(path)
    if path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(path)
        return [
            SheetGrid(
                title=sheet.name,
                rows=[
                    [sheet.cell_value(row, col) for col in range(sheet.ncols)]
                    for row in range(sheet.nrows)
                ],
            )
            for sheet in book.sheets()
        ]

    workbook = load_workbook(path, data_only=True, read_only=False)
    grids = []
    for ws in workbook.worksheets:
        rows = [
            [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
            for row in range(1, ws.max_row + 1)
        ]
        grids.append(SheetGrid(title=ws.title, rows=rows))
    return grids


def score_report_sheet(sheet: SheetGrid) -> int:
    title = sheet.title.strip().lower()
    title_boost = 0
    if title in {"eb", "endbericht"} or "endbericht" in title:
        title_boost += 100
    elif "zwb" in title:
        title_boost += 20
    elif "overview" in title:
        title_boost -= 10

    text_parts = []
    for row in range(1, min(sheet.max_row, 80) + 1):
        text_parts.extend(sheet.row_values(row, max_col=min(sheet.max_column, 40)))
    return title_boost + len(REPORT_KEYWORDS.findall(" ".join(text_parts)))


def score_header_row(sheet: SheetGrid, row: int) -> int:
    values = sheet.row_values(row, max_col=min(sheet.max_column, 80))
    keyword_hits = sum(1 for value in values if HEADER_KEYWORDS.search(value))
    has_soll = int(any("soll" in value.lower() for value in values))
    has_ist = int(any("ist" in value.lower() for value in values))
    has_percent = int(any("erfüllt" in value.lower() or "erfuellt" in value.lower() for value in values))
    return keyword_hits + has_soll + has_ist + has_percent


def detect_header_rows(sheet: SheetGrid) -> list[int]:
    scored = [(row, score_header_row(sheet, row)) for row in range(1, min(sheet.max_row, 120) + 1)]
    best_row, best_score = max(scored, key=lambda item: item[1])
    if best_score < 2:
        raise ValueError(f"Could not detect Soll/Ist header row in sheet {sheet.title}")

    header_rows = [best_row]
    next_text = " ".join(sheet.row_values(best_row + 1, max_col=min(sheet.max_column, 80)))
    if re.search(r"\bM\s*\d+\b", next_text) or score_header_row(sheet, best_row + 1) >= 2:
        header_rows.append(best_row + 1)
    return header_rows


def detect_indicator_column(sheet: SheetGrid, header_rows: list[int]) -> int:
    data_start = max(header_rows) + 1
    col_scores = []
    for col in range(1, min(sheet.max_column, 40) + 1):
        score = 0
        for row in range(data_start, min(sheet.max_row, data_start + 100) + 1):
            text = clean_text(sheet.value(row, col))
            if len(text) >= 12 and numeric_or_none(text) is None:
                score += 1
                if INDICATOR_HINTS.search(text):
                    score += 2
                if text.lower().startswith("bereich"):
                    score += 1
        col_scores.append((col, score))
    best_col, best_score = max(col_scores, key=lambda item: item[1])
    if best_score == 0:
        raise ValueError(f"Could not detect indicator label column in sheet {sheet.title}")
    return best_col


def build_column_headers(sheet: SheetGrid, header_rows: list[int]) -> dict[int, str]:
    headers = {}
    for col in range(1, sheet.max_column + 1):
        parts = [clean_text(sheet.value(row, col)) for row in header_rows]
        header = re.sub(r"\s+", " ", " ".join(part for part in parts if part)).strip()
        if header:
            headers[col] = header
    return headers


def find_local_header_rows(sheet: SheetGrid, indicator_row: int, global_header_rows: list[int]) -> list[int]:
    candidates = []
    for row in range(max(1, indicator_row - 8), indicator_row):
        score = score_header_row(sheet, row)
        if score >= 2:
            candidates.append((row, score))
    if not candidates:
        return global_header_rows

    best_row, _ = max(candidates, key=lambda item: item[1])
    header_rows = [best_row]
    if best_row + 1 < indicator_row:
        next_text = " ".join(sheet.row_values(best_row + 1, max_col=min(sheet.max_column, 80)))
        if re.search(r"\bM\s*\d+\b", next_text) or score_header_row(sheet, best_row + 1) >= 2:
            header_rows.append(best_row + 1)
    return header_rows


def detect_indicator_table(workbook_path: Path) -> dict:
    sheets = load_workbook_grids(workbook_path)
    sheets_by_score = sorted(
        [(sheet.title, score_report_sheet(sheet), sheet) for sheet in sheets],
        key=lambda item: item[1],
        reverse=True,
    )
    sheet = sheets_by_score[0][2]
    header_rows = detect_header_rows(sheet)
    indicator_col = detect_indicator_column(sheet, header_rows)
    return {
        "sheet": sheet,
        "sheet_name": sheet.title,
        "sheet_scores": [(title, score) for title, score, _ in sheets_by_score],
        "header_rows": header_rows,
        "indicator_col": indicator_col,
        "data_start": max(header_rows) + 1,
    }


def column_match_score(
    models: ExtractionModels,
    query: str,
    header: str,
    required_terms: list[str] | None = None,
    forbidden_terms: list[str] | None = None,
) -> float:
    base_score = float(models.cosine_scores(query, [header])[0])
    text = normalize_for_match(header)
    score = base_score
    for term in required_terms or []:
        if term in text:
            score += 0.12
    for term in forbidden_terms or []:
        if term in text:
            score -= 0.18
    return score


def best_column(
    models: ExtractionModels,
    headers: dict[int, str],
    query: str,
    required_terms: list[str] | None = None,
    forbidden_terms: list[str] | None = None,
    min_score: float = 0.35,
) -> tuple[int, str, float]:
    scored = []
    for col, header in headers.items():
        score = column_match_score(
            models,
            query,
            header,
            required_terms=required_terms,
            forbidden_terms=forbidden_terms,
        )
        scored.append((col, header, score))
    best = max(scored, key=lambda item: item[2])
    if best[2] < min_score:
        raise ValueError(f"Could not match column for query: {query}")
    return best


def extract_indicator_kpi_dynamic(
    workbook_path: Path,
    models: ExtractionModels,
    indicator_query: str = "Anzahl der Projektteilnehmenden gesamt",
) -> dict:
    table = detect_indicator_table(workbook_path)
    sheet = table["sheet"]
    indicator_col = table["indicator_col"]

    row_candidates = []
    for row in range(table["data_start"], sheet.max_row + 1):
        label = clean_text(sheet.value(row, indicator_col))
        if label and len(label) > 6 and not label.lower().startswith("bereich"):
            row_candidates.append((row, label))

    best_row = models.best_match(indicator_query, [label for _, label in row_candidates], min_score=0.45)
    if best_row is None:
        raise ValueError(f"Could not find indicator row for query: {indicator_query}")

    row_idx, row_score = best_row
    indicator_row, matched_indicator = row_candidates[row_idx]
    local_header_rows = find_local_header_rows(sheet, indicator_row, table["header_rows"])
    headers = build_column_headers(sheet, local_header_rows)

    soll_col, soll_header, soll_score = best_column(
        models,
        headers,
        "Anzahl Soll laut Vertrag",
        required_terms=["soll"],
        forbidden_terms=["maßnahme", "massnahme", " m 1", " m1"],
    )
    ist_col, ist_header, ist_score = best_column(
        models,
        headers,
        "Anzahl Ist Endbericht gesamt erreicht",
        required_terms=["ist"],
        forbidden_terms=["maßnahme", "massnahme", " m 1", " m1"],
    )
    percent_col, percent_header, percent_score = best_column(
        models,
        headers,
        "Erfüllt in Prozent Zielerreichung",
        required_terms=["erfüllt"],
    )

    soll = numeric_or_none(sheet.value(indicator_row, soll_col))
    ist = numeric_or_none(sheet.value(indicator_row, ist_col))
    fulfillment = numeric_or_none(sheet.value(indicator_row, percent_col))
    if fulfillment is not None and fulfillment <= 2:
        fulfillment_percent = fulfillment * 100
    else:
        fulfillment_percent = fulfillment

    return {
        "indikator": matched_indicator,
        "soll": format_number(soll),
        "ist": format_number(ist),
        "erfuellung_prozent": round(fulfillment_percent, 1) if fulfillment_percent is not None else None,
        "confidence": round(min(row_score, soll_score, ist_score, percent_score), 3),
        "evidence": {
            "source_file": str(workbook_path),
            "sheet": table["sheet_name"],
            "global_header_rows": table["header_rows"],
            "local_header_rows": local_header_rows,
            "indicator_col": indicator_col,
            "indicator_row": indicator_row,
            "matched_indicator": matched_indicator,
            "soll_header": soll_header,
            "soll_cell": sheet.cell(indicator_row, soll_col).coordinate,
            "ist_header": ist_header,
            "ist_cell": sheet.cell(indicator_row, ist_col).coordinate,
            "percent_header": percent_header,
            "percent_cell": sheet.cell(indicator_row, percent_col).coordinate,
        },
    }


def non_empty_cells(sheet: SheetGrid, max_row: int = 50, max_col: int = 25) -> list[dict]:
    cells = []
    for row in range(1, min(max_row, sheet.max_row) + 1):
        for col in range(1, min(max_col, sheet.max_column) + 1):
            cell = sheet.cell(row, col)
            text = cell.text
            if text:
                cells.append(
                    {
                        "sheet": sheet.title,
                        "row": row,
                        "col": col,
                        "coordinate": cell.coordinate,
                        "text": text,
                        "value": cell.value,
                    }
                )
    return cells


def nearest_right_value(sheet: SheetGrid, row: int, col: int, max_steps: int = 10):
    for next_col in range(col + 1, min(sheet.max_column, col + max_steps) + 1):
        cell = sheet.cell(row, next_col)
        if cell.text:
            return cell.value, cell.coordinate
    return None, None


def extract_key_value_semantic_any_sheet(
    workbook_path: Path,
    models: ExtractionModels,
    german_query: str,
    max_row: int = 50,
    max_col: int = 25,
    min_score: float = 0.48,
) -> dict:
    sheets = load_workbook_grids(workbook_path)
    all_cells = []
    sheet_by_title = {sheet.title: sheet for sheet in sheets}
    for sheet in sheets:
        all_cells.extend(non_empty_cells(sheet, max_row=max_row, max_col=max_col))

    best = models.best_match(german_query, [cell["text"] for cell in all_cells], min_score=min_score)
    if best is None:
        return {"value": None, "confidence": 0.0, "evidence": None}

    idx, score = best
    label_cell = all_cells[idx]
    sheet = sheet_by_title[label_cell["sheet"]]
    value, value_coordinate = nearest_right_value(sheet, label_cell["row"], label_cell["col"])
    return {
        "value": clean_text(value),
        "confidence": round(score, 3),
        "evidence": {
            "source_file": str(workbook_path),
            "sheet": label_cell["sheet"],
            "matched_label": label_cell["text"],
            "label_cell": label_cell["coordinate"],
            "value_cell": value_coordinate,
        },
    }


def extract_project_metadata(workbook_path: Path, models: ExtractionModels) -> dict:
    excel_fields = {
        "projekttraeger": "Projektträger oder Trägerorganisation",
        "projekttitel": "Projekttitel oder Projektname",
        "projektnummer": "Projektnummer oder Geschäftszeichen",
        "laufzeit_beginn": "Laufzeit Beginn oder Projektstart",
        "laufzeit_ende": "Laufzeit Ende oder Projektende",
        "hauptprojektstandort": "Hauptprojektstandort oder Bundesland des Projekts",
    }
    project_meta = {
        field: extract_key_value_semantic_any_sheet(workbook_path, models, query)
        for field, query in excel_fields.items()
    }
    project_meta["laufzeit_beginn"]["value"] = parse_excel_date(project_meta["laufzeit_beginn"]["value"])
    project_meta["laufzeit_ende"]["value"] = parse_excel_date(project_meta["laufzeit_ende"]["value"])
    return project_meta
